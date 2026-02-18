import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows

def generate_excel(safety_df, recon_df, output_file="Query_Log.xlsx"):
    """
    Generates an Excel report with highlighted queries.
    """
    print(f"[INFO] Generating report: {output_file}...")
    
    wb = Workbook()
    
    # --- Tab 1: Safety Queries ---
    ws1 = wb.active
    ws1.title = "Safety Queries"
    
    # Convert dataframe to rows with header
    # Include relevant columns only if too many
    # Convert dataframe to rows with header
    # Include relevant columns only if too many
    cols = ["USUBJID", "INITIALS", "BRTHDTC", "SVSTDTC", "VSSTRESN_TEMP", "VSSTRESN_HR", "Raw_RR", "WBC", "Query_Text"]
    # Check if cols exist (some might be missing if no merges happened correctly or empty df)
    existing_cols = [c for c in cols if c in safety_df.columns]
    
    if not safety_df.empty:
        rows = dataframe_to_rows(safety_df[existing_cols], index=False, header=True)
    else:
        rows = [existing_cols] # just header
        
    for r_idx, row in enumerate(rows, 1):
        for c_idx, value in enumerate(row, 1):
            cell = ws1.cell(row=r_idx, column=c_idx, value=value)
            
            # Highlight header
            if r_idx == 1:
                cell.fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
            else:
                # Highlight rows RED
                cell.fill = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")

    # --- Tab 2: Recon Queries ---
    ws2 = wb.create_sheet(title="Recon Queries")
    
    cols_recon = ["USUBJID", "INITIALS", "BRTHDTC", "SVSTDTC", "Blood_Draw_Performed", "SampleID", "Query_Text"]
    existing_cols_recon = [c for c in cols_recon if c in recon_df.columns]
    
    if not recon_df.empty:
        rows_recon = dataframe_to_rows(recon_df[existing_cols_recon], index=False, header=True)
    else:
        rows_recon = [existing_cols_recon]
        
    for r_idx, row in enumerate(rows_recon, 1):
        for c_idx, value in enumerate(row, 1):
            cell = ws2.cell(row=r_idx, column=c_idx, value=value)
            
            # Highlight header
            if r_idx == 1:
                cell.fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
            else:
                # Highlight rows YELLOW
                cell.fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")

    wb.save(output_file)
    print(f"[INFO] Report generated successfully: {output_file}")

if __name__ == "__main__":
    # Test stub
    pass
